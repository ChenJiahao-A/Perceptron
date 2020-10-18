import openpyxl
import random
import matplotlib.pyplot as plt
import numpy as np


def generate(workbook, worksheet, dimension):
    row = 2
    x = np.zeros([dimension], dtype=float, order='C')
    while row < 200:
        sign1 = 50
        sign2 = 60
        signal = -1
        for i in range(dimension):
            x[i] = random.randint(-100, 100)
            sign1 = sign1 + signal * x[i]
            sign2 = sign2 + signal * x[i]
            worksheet.cell(row, i+2).value = x[i]
            signal = -signal
        if sign1 > 0:
            mark = 1
        elif sign2 < 0:
            mark = -1
        else:
            continue
        worksheet.cell(row, dimension+2).value = mark
        row = row + 1
    workbook.save("Perceptron.xlsx")


def draw_dot(graph, worksheet, dimension):
    c = [[[] for j in range(dimension)]for i in range(2)]

    for raw in range(2, worksheet.max_row + 1):
        if worksheet.cell(raw, dimension+2).value == 1:
            for i in range(dimension):
                c[0][i].append(worksheet.cell(raw, i+2).value)
        else:
            for i in range(dimension):
                c[1][i].append(worksheet.cell(raw, i+2).value)
    if dimension == 3:
        graph.scatter(c[0][0], c[0][1], c[0][2], c='r')
        graph.scatter(c[1][0], c[1][1], c[1][2], c='g')
        graph.set_xlabel('x1')
        graph.set_ylabel('x2')
        graph.set_zlabel('x3')
    elif dimension == 2:
        graph.scatter(c[0][0], c[0][1], c='r')
        graph.scatter(c[1][0], c[1][1], c='g')
        graph.set_xlabel('x1')
        graph.set_ylabel('x2')


def draw_original(graph, worksheet, dimension):
    w = np.ones([dimension], dtype=float, order='C')
    x = np.zeros([dimension], dtype=float, order='C')
    judge = 0
    b = 1
    eta = 1
    i = 2
    while i < worksheet.max_row + 1:
        for j in range(dimension):
            x[j] = worksheet.cell(i, j+2).value
            judge = judge + x[j]*w[j]
        y = worksheet.cell(i, dimension+2).value
        judge = y*(judge + b)
        while judge <= 0:
            judge = 0
            for k in range(dimension):
                w[k] = w[k] + eta * x[k] * y
                judge = judge + x[k]*w[k]
            b = b + eta * y
            judge = y * (judge + b)
            i = 2
        judge = 0
        i = i + 1

    if dimension == 2:
        x = np.arange(-100, 100)
        y = -(w[0]*x + b)/w[1]
        graph.plot(x, y)
    elif dimension == 3:
        x1 = np.arange(-100, 100)
        x2 = np.arange(-100, 100)
        x1, x2 = np.meshgrid(x1, x2)
        y = -(x1*w[0] + x2*w[1] + b)/w[2]
        graph.plot_surface(x1, x2, y)


def draw_dual(graph, worksheet, dimension):
    gram = np.zeros([worksheet.max_row-1, worksheet.max_row-1], dtype=float, order='C')
    for i in range(worksheet.max_row-1):
        for j in range(i, worksheet.max_row-1):
            for k in range(dimension):
                gram[i][j] = gram[i][j] + worksheet.cell(i+2, k+2).value*worksheet.cell(j+2, k+2).value
            gram[j][i] = gram[i][j]
    a = np.zeros([worksheet.max_row - 1], dtype=float, order='C')
    b = 0
    eta = 0.1
    judge = 0
    i = 0
    while i < worksheet.max_row - 1:
        temp = 1
        y = worksheet.cell(i + 2, dimension + 2).value
        while judge <= 0:
            judge = 0
            for j in range(0, worksheet.max_row-1):
                yj = worksheet.cell(j+2, dimension+2).value
                judge = judge + a[j]*yj*gram[i][j]
            judge = y*(judge + b)
            if judge <= 0:
                a[i] = a[i] + eta
                b = b + eta*y
                temp = 0
        judge = 0
        i = i + 1
        if temp == 0:
            i = 0
    w = np.zeros([dimension], dtype=float, order='c')
    for i in range(worksheet.max_row-1):
        y = worksheet.cell(i + 2, dimension + 2).value
        for j in range(dimension):
            x = worksheet.cell(i+2, j+2).value
            w[j] = w[j] + a[i]*x*y

    if dimension == 2:
        x = np.arange(-100, 100)
        y = -(w[0]*x + b)/w[1]
        graph.plot(x, y)
    elif dimension == 3:
        x1 = np.arange(-100, 100)
        x2 = np.arange(-100, 100)
        x1, x2 = np.meshgrid(x1, x2)
        y = -(x1*w[0] + x2*w[1] + b)/w[2]
        graph.plot_surface(x1, x2, y)


def draw_2dimension():
    wb = openpyxl.load_workbook("Perceptron.xlsx")
    ws = wb["two-dimension"]
    fig = plt.figure()
    gra = fig.add_subplot(111)
    generate(wb, ws, 2)
    draw_dot(gra, ws, 2)

    #  draw_dual(gra, ws, 2)
    draw_original(gra, ws, 2)

    plt.show()


def draw_3dimension():
    wb = openpyxl.load_workbook("Perceptron.xlsx")
    ws = wb["three-dimension"]
    generate(wb, ws, 3)
    fig = plt.figure()
    gra = fig.add_subplot(111, projection='3d')
    draw_dot(gra, ws, 3)

    draw_original(gra, ws, 3)
    #  draw_dual(gra, ws, 3)
    plt.show()


draw_2dimension()
draw_3dimension()
